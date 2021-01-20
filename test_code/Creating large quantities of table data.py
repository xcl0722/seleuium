
import openpyxl
import random
import datetime
#生成一个 Workbook 的实例化对象，wb即代表一个工作簿（一个 Excel 文件）
wb = openpyxl.Workbook()
# 获取活跃的工作表，ws代表wb(工作簿)的一个工作表
ws = wb.active
#更改工作表ws的title
ws.title = 'test_sheet1'
# #对ws的单个单元格传入数据
# ws['A1'] = 'SiemensSensorCode'
# ws['B1'] = 'Time'
# ws['C1'] = 'Value'
# data = {
#     '中国':'北京',
#     '韩国':'首尔',
#     '日本':'东京',
#     '泰国':'曼谷',
#     '马来西亚':'吉隆坡',
#     '越南':'河内',
#     '朝鲜':'平壤',
#     '印度':'新德里'
#     }
# data_excel = []
# #将字典中的每对数据（键，值）以列表形式传入data_excel列表
# for each in data:
#     data_excel.append([each, data[each]])
# #将data_excel列表内的内容存入工作表
# for each in data_excel:
#     ws.append(each)
# #注意：上述两个append方法是意义完全不同的两个方法
start = '05/01/2020 09:00:00'
end = '07/14/2020 08:59:59'
datestart = datetime.datetime.strptime(start, '%m/%d/%Y %H:%M:%S')
dateend = datetime.datetime.strptime(end, '%m/%d/%Y %H:%M:%S')
# 生成按照一定规律递增的时间列表
data_list = list()
while datestart <= dateend:
    data_list.append(datestart.strftime('%m/%d/%Y %H:%M:%S'))
    datestart += datetime.timedelta(seconds=30)
"""添加字段"""
ws.cell(1, 1, 'SiemensSensorCode')
ws.cell(1, 2, 'Time')
ws.cell(1, 3, 'Value')

"""写入编号字段数据"""
for id in range(60):
    ws.cell(id + 1, 1, 'xcl_XNDF-666666-1-SENSOR95')
# 依次取出列表里的时间值
    ws.cell(id + 1, 2, data_list[id])
    ws.cell(id + 1, 3, round(random.uniform(20, 90), 0))
wb.save(r'C:\Users\huarui\Desktop\test1.xls')